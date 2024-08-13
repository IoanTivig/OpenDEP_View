from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Border, Side
from openpyxl.utils import get_column_letter


def save_scatter_to_excel(file, scatter_data):
    wb = Workbook()
    ws = wb.active

    ## Excel top-table setup ##
    ws['A1'] = 'Frequency (Hz)'
    ws['B1'] = 'Experimental CM factor'
    ws['C1'] = 'Experimental CM Factor errors'

    auto_stretch_columns(ws)
    set_background_color(ws, top=1, bottom=1000, left=1, right=100, color="00EBF1DE")
    set_background_color(ws, top=1, bottom=1, left=1, right=3, color='00C4D79B')
    set_border(ws, top=1, bottom=1, left=1, right=3, color='000000', border_style='thick')

    ## Excel bottom-table setup ##
    print(scatter_data['frequencies'])
    for a in range(len(scatter_data['frequencies'])):
        ws.cell(row=a+2, column=1).value = scatter_data['frequencies'][a]
        ws.cell(row=a+2, column=2).value = scatter_data['recm_values'][a]
        ws.cell(row=a+2, column=3).value = scatter_data['recm_errors'][a]

    set_border(ws, top=2, bottom=len(scatter_data['frequencies']) + 1, left=1, right=3, color='000000', border_style='double')
    set_tabel_background(ws, top=2, bottom=len(scatter_data['frequencies']) + 1, left=1, right=3, color1='00C0C0C0', color2='00FFFFFF')

    wb.save(filename=file)


def load_scatter_from_excel(file):
    wb = load_workbook(filename=file)
    ws = wb.active

    scatter_data = {
        'frequencies': [],
        'recm_values': [],
        'recm_errors': []
    }

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=3):
        if row[0].value is None:
            break
        scatter_data['frequencies'].append(row[0].value)
        scatter_data['recm_values'].append(row[1].value)
        scatter_data['recm_errors'].append(row[2].value)

    # Check if all values are numbers and if the list is not empty and all lists have the same length
    if all(isinstance(x, (int, float)) for x in scatter_data['frequencies']) and \
            all(isinstance(x, (int, float)) for x in scatter_data['recm_values']) and \
            all(isinstance(x, (int, float)) for x in scatter_data['recm_errors']) and \
            len(scatter_data['frequencies']) > 0 and \
            len(scatter_data['frequencies']) == len(scatter_data['recm_values']) == len(scatter_data['recm_errors']):
        return scatter_data

    else:
        return None


def load_curve_from_excel(file):
    wb = load_workbook(filename=file)
    ws = wb.active
    parameters = {}
    if ws.cell(row=2, column=12).value == "sish":
        parameters = {"buffer_perm": ws.cell(row=7, column=12).value,
                      "buffer_cond": ws.cell(row=6, column=12).value,
                      "core_perm": ws.cell(row=15, column=12).value,
                      "core_cond": ws.cell(row=14, column=12).value,
                      "core_radius": ws.cell(row=8, column=12).value,
                      "1st_shell_perm": ws.cell(row=11, column=12).value,
                      "1st_shell_cond": ws.cell(row=10, column=12).value,
                      "1st_shell_thick": ws.cell(row=9, column=12).value,
                      "2nd_shell_perm": 10,
                      "2nd_shell_cond": 0.00001,
                      "2nd_shell_thick": 6.0,
                      "electric_field": 1.0,
                      "1st_cross_over": {"homogenous": 0.0,
                                         "single_shell": 0.0,
                                         "two_shell": 0.0},
                      "2nd_cross_over": {"homogenous": 0.0,
                                         "single_shell": 0.0,
                                         "two_shell": 0.0}
                      }
        model = 1

    elif ws.cell(row=2, column=12).value == "hopa":
        parameters = {"buffer_perm": ws.cell(row=7, column=12).value,
                      "buffer_cond": ws.cell(row=6, column=12).value,
                      "core_perm": ws.cell(row=9, column=12).value,
                      "core_cond": ws.cell(row=8, column=12).value,
                      "core_radius": ws.cell(row=10, column=12).value,
                      "1st_shell_perm": 40,
                      "1st_shell_cond": 0.00001,
                      "1st_shell_thick": 6.0,
                      "2nd_shell_perm": 10,
                      "2nd_shell_cond": 0.00001,
                      "2nd_shell_thick": 6.0,
                      "electric_field": 1.0,
                      "1st_cross_over": {"homogenous": 0.0,
                                         "single_shell": 0.0,
                                         "two_shell": 0.0},
                      "2nd_cross_over": {"homogenous": 0.0,
                                         "single_shell": 0.0,
                                         "two_shell": 0.0}
                      }
        model = 0

    # test for all parameters to be numbers or dictionaries and not empty dict
    if len(parameters) > 0 and all(isinstance(x, (int, float, dict)) for x in parameters.values()) and \
            all(isinstance(x, (int, float)) for x in parameters['1st_cross_over'].values()) and \
            all(isinstance(x, (int, float)) for x in parameters['2nd_cross_over'].values()):
        print(parameters, model)
        return parameters, model
    else:
        return None, None


# Setting borders to a group of cells #
def set_border(ws, top, bottom, left, right, color, border_style):
    for i in range(top, bottom):
        border1 = Border(left=Side(border_style=border_style, color=color))
        ws.cell(row=i, column=left).border = border1
        border2 = Border(right=Side(border_style=border_style, color=color))
        ws.cell(row=i, column=right).border = border2

    for i in range(left, right):
        border1 = Border(top=Side(border_style=border_style, color=color))
        ws.cell(row=top, column=i).border = border1
        border2 = Border(bottom=Side(border_style=border_style, color=color))
        ws.cell(row=bottom, column=i).border = border2

    border = Border(left=Side(border_style=border_style, color=color), top=Side(border_style=border_style, color=color))
    ws.cell(row=top, column=left).border = border
    border = Border(left=Side(border_style=border_style, color=color),
                    bottom=Side(border_style=border_style, color=color))
    ws.cell(row=bottom, column=left).border = border
    border = Border(top=Side(border_style=border_style, color=color),
                    right=Side(border_style=border_style, color=color))
    ws.cell(row=top, column=right).border = border
    border = Border(bottom=Side(border_style=border_style, color=color),
                    right=Side(border_style=border_style, color=color))
    ws.cell(row=bottom, column=right).border = border


def set_background_color(ws, top, bottom, left, right, color):
    fill = PatternFill("solid", fgColor=color)
    for i in range(top, bottom + 1):
        for j in range(left, right + 1):
            ws.cell(row=i, column=j).fill = fill


def set_tabel_background(ws, top, bottom, left, right, color1, color2):
    for i in range(top, bottom + 1):
        if (i % 2) == 0:
            set_background_color(ws, top=i, bottom=i, left=left, right=right, color=color1)
        else:
            set_background_color(ws, top=i, bottom=i, left=left, right=right, color=color2)


def auto_stretch_columns(ws):
    for col in ws.columns:
        max_length = 0
        column = get_column_letter(col[0].column)  # Get the column name
        # Since Openpyxl 2.6, the column name is  ".column_letter" as .column became the column number (1-based)
        for cell in col:
            try:  # Necessary to avoid error on empty cells
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1
        ws.column_dimensions[column].width = adjusted_width
